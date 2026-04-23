"""
OCR extraction using Gemini 2.5 Flash.

Supports:
  - PDF  → converts each page to a PNG via PyMuPDF, then sends to Gemini Vision
  - Images (JPG / PNG / WEBP) → sent directly

Returns raw extracted text (preserving layout as much as possible).
"""

import io
import base64
import logging
from pathlib import Path

import google.generativeai as genai
from PIL import Image

from app.core.config import settings

logger = logging.getLogger(__name__)

# NOTE: configure() được gọi lại mỗi khi dùng để luôn đọc key mới nhất
SUPPORTED_MIME: dict[str, str] = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
}

_OCR_PROMPT = (
    "Extract ALL text from this document image. "
    "Preserve the original layout including tables, headers, line breaks, and columns. "
    "Return only the extracted text — no commentary."
)


def _model() -> genai.GenerativeModel:
    key = settings.GEMINI_API_KEY
    logger.info("Gemini key prefix: %s...", key[:12] if key else "(empty)")
    genai.configure(api_key=key)
    return genai.GenerativeModel(settings.GEMINI_MODEL)


def _pil_to_part(pil_image: Image.Image) -> Image.Image:
    """Return a PIL image that the Gemini SDK accepts directly."""
    return pil_image


async def extract_text_from_file(file_path: str) -> str:
    """
    Main entry point.  Given a local file path, return a single string
    containing all OCR-extracted text.
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".pdf":
        return _extract_pdf(file_path)
    elif ext in SUPPORTED_MIME:
        return _extract_image(file_path)
    elif ext == ".docx":
        return _extract_docx(file_path)
    elif ext == ".xlsx":
        return _extract_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file extension: {ext!r}")


# ─── PDF ─────────────────────────────────────────────────────────────────────

def _extract_pdf(file_path: str) -> str:
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise RuntimeError("PyMuPDF is required for PDF support: pip install PyMuPDF")

    model = _model()
    page_texts: list[str] = []

    doc = fitz.open(file_path)
    total = len(doc)
    logger.info("Extracting text from PDF '%s' (%d pages)", file_path, total)

    for page_num in range(total):
        page = doc[page_num]
        # 2× zoom gives sharper images → better OCR accuracy
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
        img_bytes = pix.tobytes("png")
        pil_img = Image.open(io.BytesIO(img_bytes))

        response = model.generate_content([_OCR_PROMPT, pil_img])
        page_texts.append(f"[Trang {page_num + 1}/{total}]\n{response.text}")
        logger.debug("  Page %d extracted (%d chars)", page_num + 1, len(response.text))

    doc.close()
    return "\n\n".join(page_texts)


# ─── Image ────────────────────────────────────────────────────────────────────

def _extract_image(file_path: str) -> str:
    model = _model()
    pil_img = Image.open(file_path)
    logger.info("Extracting text from image '%s'", file_path)
    response = model.generate_content([_OCR_PROMPT, pil_img])
    return response.text


# ─── DOCX ─────────────────────────────────────────────────────────────────────

def _extract_docx(file_path: str) -> str:
    """
    Extract text from a .docx file using python-docx.
    Paragraphs are returned as-is; tables are rendered as pipe-separated rows.
    The resulting text is then sent to Gemini for structured extraction (same
    pipeline as PDF/image — only the OCR step is replaced by direct parsing).
    """
    try:
        from docx import Document as DocxDocument
        from docx.table import Table
        from docx.text.paragraph import Paragraph
        from docx.oxml.ns import qn
    except ImportError:
        raise RuntimeError("python-docx is required: pip install python-docx")

    logger.info("Extracting text from DOCX '%s'", file_path)
    doc = DocxDocument(file_path)
    parts: list[str] = []

    def _table_to_text(tbl: Table) -> str:
        rows = []
        for row in tbl.rows:
            cells = [cell.text.strip().replace("\n", " ") for cell in row.cells]
            rows.append(" | ".join(cells))
        return "\n".join(rows)

    # Iterate body elements in document order (paragraphs + tables interleaved)
    for block in doc.element.body:
        tag = block.tag.split("}")[-1] if "}" in block.tag else block.tag
        if tag == "p":
            para = Paragraph(block, doc)
            text = para.text.strip()
            if text:
                parts.append(text)
        elif tag == "tbl":
            tbl = Table(block, doc)
            parts.append(_table_to_text(tbl))

    raw_text = "\n".join(parts)
    logger.debug("DOCX extracted %d chars", len(raw_text))
    return raw_text


# ─── XLSX ─────────────────────────────────────────────────────────────────────

def _extract_xlsx(file_path: str) -> str:
    """
    Extract data from a .xlsx file using openpyxl.
    Each sheet is rendered as a markdown-style table separated by a sheet header.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    logger.info("Extracting text from XLSX '%s'", file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets_text: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            cells = [str(v) if v is not None else "" for v in row]
            rows.append("\t".join(cells))
        if rows:
            sheets_text.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))

    wb.close()
    raw_text = "\n\n".join(sheets_text)
    logger.debug("XLSX extracted %d chars", len(raw_text))
    return raw_text
